from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Series, Reference
wb = Workbook()
ws = wb.active
person = [["id", "姓名", "年龄"], ["1", "小李",23 ], ["2", "小刘", 73], ["3", "小万", 24]]

for row in person:
    ws.append(row)
ft = Font(bold=True)
for row in ws["A1:C1"]:
 for cell in row:
  cell.font = ft

newwb = wb.create_sheet('newsheet')
for i in range (1,10):
    newwb.cell(10,i,value='test')


#      制作表格
chart = BarChart()
chart.type = "col"
chart.title = "Tree Height"
chart.y_axis.title = 'Height (cm)'
chart.x_axis.title = 'Tree Type'
chart.legend = None

data = Reference(ws, min_col=3, min_row=2, max_row=4, max_col=3)
categories = Reference(ws, min_col=1, min_row=2, max_row=4, max_col=1)

chart.add_data(data)
chart.set_categories(categories)

ws.add_chart(chart, "E1")
wb.save('person.xlsx')


